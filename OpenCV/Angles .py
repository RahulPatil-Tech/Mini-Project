import cv2
import numpy as np
import openpyxl

cap = cv2.VideoCapture(0)

# Parameters for edge detection
threshold1 = 50
threshold2 = 150

# Parameters for line detection
rho = 1
theta = np.pi/180
threshold = 50
min_line_length = 100
max_line_gap = 10

# Create a new Excel workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Angle Measurements'
ws['A1'] = 'Frame'
ws['B1'] = 'Angle (degrees)'
row = 2

while True:
    # Read a frame from the camera
    ret, frame = cap.read()

    # If we failed to read a frame, abort
    if not ret:
        print("Failed to read frame from camera.")
        break

    # Convert the frame to grayscale for edge detection
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # Perform edge detection on the grayscale image
    edges = cv2.Canny(gray, threshold1, threshold2, apertureSize=3)

    # Detect lines in the image using the HoughLinesP method
    lines = cv2.HoughLinesP(edges, rho, theta, threshold, minLineLength=min_line_length, maxLineGap=max_line_gap)

    # Only process if we detected at least two lines
    if lines is not None and len(lines) >= 2:
        # Draw the lines on the frame
        for line in lines:
            x1, y1, x2, y2 = line[0]
            cv2.line(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)

        # Calculate the angle between the two lines
        x1, y1, x2, y2 = lines[0][0]
        line1_angle = np.arctan2(y2 - y1, x2 - x1)

        x1, y1, x2, y2 = lines[1][0]
        line2_angle = np.arctan2(y2 - y1, x2 - x1)

        angle = np.abs(np.degrees(line1_angle - line2_angle))

        # Display the angle on the frame
        cv2.putText(frame, f"Angle: {angle:.2f} deg", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

        # Write the angle to the Excel sheet
        ws.cell(row=row, column=1).value = row - 1
        ws.cell(row=row, column=2).value = angle
        row += 1

    # Display the resulting frame
    cv2.imshow('frame', frame)

    # Check for the 'q' key and exit if pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Save the Excel workbook
wb.save('angle_measurements.xlsx')

# Release the capture and destroy all windows
cap.release()
cv2.destroyAllWindows()